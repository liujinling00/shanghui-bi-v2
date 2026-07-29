# -*- coding: utf-8 -*-
"""
Excel 转 JSON 工具
将商会会员 Excel 表格转换为 data.json 格式

使用方法:
    python excel_to_json.py 会员数据.xlsx
    python excel_to_json.py 会员数据.xlsx --output data.json

Excel 表头要求（第一行）:
    姓名 | 性别 | 年龄段 | 政治面貌 | 所属行业 | 区域 | 区县 | 会内职位 | 入会年份 | 会费状态 | 所属企业 | 好人好事

安装依赖:
    pip install openpyxl
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("错误：需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


def excel_to_json(excel_path, output_path="data.json"):
    """读取 Excel 文件并转换为 data.json 格式"""

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # 读取表头
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value else ""
        headers.append(val)

    print(f"检测到表头: {headers}")

    # 字段映射（支持中英文表头）
    field_map = {
        "姓名": "name", "name": "name",
        "性别": "gender", "gender": "gender",
        "年龄段": "ageGroup", "年龄": "ageGroup", "age": "ageGroup",
        "政治面貌": "politicalStatus", "political": "politicalStatus",
        "所属行业": "industry", "行业": "industry", "industry": "industry",
        "区域": "region", "region": "region",
        "区县": "district", "district": "district",
        "会内职位": "position", "职位": "position", "position": "position",
        "入会年份": "joinYear", "入会时间": "joinYear", "joinYear": "joinYear",
        "会费状态": "membershipStatus", "状态": "membershipStatus", "status": "membershipStatus",
        "所属企业": "company", "企业": "company", "company": "company",
        "好人好事": "goodDeeds", "goodDeeds": "goodDeeds",
    }

    # 建立列索引映射
    col_map = {}
    for i, h in enumerate(headers):
        key = field_map.get(h)
        if key:
            col_map[key] = i

    required_fields = ["name", "gender", "ageGroup", "politicalStatus", "industry", "region", "district", "position", "joinYear", "membershipStatus"]
    missing = [f for f in required_fields if f not in col_map]
    if missing:
        print(f"\n警告：以下必填字段未在 Excel 中找到: {missing}")
        print("请检查 Excel 表头是否包含这些列。")
        print("支持的表头名称：姓名、性别、年龄段、政治面貌、所属行业、区域、区县、会内职位、入会年份、会费状态、所属企业")
        proceed = input("是否继续？(y/n): ")
        if proceed.lower() != 'y':
            sys.exit(0)

    # 读取数据行
    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        member = {"id": len(members) + 1}
        for field, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                val = ""
            if field == "joinYear":
                try:
                    val = int(val) if val else 0
                except (ValueError, TypeError):
                    val = 0
            elif field == "goodDeeds":
                try:
                    val = int(val) if val else 0
                except (ValueError, TypeError):
                    val = 0
            else:
                val = str(val).strip()
            member[field] = val

        members.append(member)

    # 生成数据
    data = {
        "meta": {
            "title": "沈阳市年轻一代民营企业家商会",
            "subtitle": "BI 智能看板",
            "lastUpdate": datetime.now().strftime("%Y-%m-%d"),
            "version": "2.0",
            "description": "商会会员数据分析与可视化平台"
        },
        "members": members
    }

    # 写入文件
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n转换成功!")
    print(f"   总记录数: {len(members)}")
    print(f"   输出文件: {os.path.abspath(output_path)}")

    # 统计信息
    if members:
        male = sum(1 for m in members if m.get("gender") == "男")
        female = sum(1 for m in members if m.get("gender") == "女")
        active = sum(1 for m in members if m.get("membershipStatus") == "在籍")
        print(f"   男性: {male}, 女性: {female}")
        print(f"   在籍: {active}, 未缴费: {len(members) - active}")

    return data


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法: python excel_to_json.py <Excel文件路径> [--output 输出文件路径]")
        print("示例: python excel_to_json.py 会员数据.xlsx")
        print("示例: python excel_to_json.py 会员数据.xlsx --output data.json")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_path = "data.json"

    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    if not os.path.exists(excel_path):
        print(f"错误：文件不存在: {excel_path}")
        sys.exit(1)

    excel_to_json(excel_path, output_path)
